import sheet
from openpyxl import load_workbook
from openpyxl.packaging import workbook

workbook = load_workbook("tmp/Еженедельные отчёты .xls")
sheet = workbook.active
print(sheet.cell(row=2, column=2).value)







